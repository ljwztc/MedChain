import pandas as pd
import re
import numpy as np
from openpyxl import load_workbook

def establish_room_structure(room_data_path):
    departments = {}
    with open(room_data_path, 'r', encoding='utf-8') as file:
        current_department = None
        for line in file:
            line = line.strip()
            if line.endswith(':'):  # 判断是否是一级科室
                current_department = line[:-1]  # 去掉冒号
                departments[current_department] = []
            elif line.startswith('-'):  # 判断是否是二级科室
                sub_department = line[2:].strip()
                if current_department:
                    departments[current_department].append(sub_department)
    # # 打印生成的字典
    # for dept, sub_depts in departments.items():
    #     print(f"{dept}: {', '.join(sub_depts)}")
    return departments

def establish_db_excel_sheet(db_excel_ptah, room_list, columns):
    # 创建一个ExcelWriter对象
    with pd.ExcelWriter(db_excel_ptah, engine='xlsxwriter') as writer:
        # 创建带有指定列名的空DataFrame
        df = pd.DataFrame(columns=columns)
        # 写入到每个sheet中
        df.to_excel(writer, index=False)

def extract_age_and_sex(string):
    # 使用正则表达式匹配年龄
    match_age = re.search(r'(\d+)岁', string)
    match_sex = re.search(r'(男|女)', string)

    if match_age:
        age = match_age.group(1)  # 提取匹配到的年龄
        age = str(age) + "岁"
    else:
        age = None
    if match_sex:
        gender = match_sex.group(1)  # 提取匹配到的性别
        gender = str(gender)
    else:
        gender = None
    result = (age, gender)
    return result


def write_excel(db_excel_ptah, case_room, case_info):
    df = pd.DataFrame({
        '年龄': [case_info[0]],
        '性别': [case_info[1]],
        '主诉': [case_info[2]],
        '症状': [case_info[3]],
        '既往史': [case_info[4]],
        '体格检查': [case_info[5]],
        '辅助检查': [case_info[6]],
        '影像报告': [case_info[7]],
        '诊断结果': [case_info[8]],
        '治疗项目': [case_info[9]],
        '一级科室': [case_info[10]],
        '二级科室': [case_info[11]]
    })

    # 加载工作簿
    book = load_workbook(db_excel_ptah)
    # 选择工作表
    sheet = book.active
    # 获取当前最大行数
    start_row = sheet.max_row

    with pd.ExcelWriter(db_excel_ptah, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, header=False, startrow=start_row)


def extract_info(case_message):
    # 提取科室
    try:
        case_room = case_message.get("tags", {}).get("科室", [None])[0]
    except Exception as e:
        case_room = None
        print(f"提取科室时发生错误: {e}")

    # 提取主诉
    try:
        description = ",".join(case_message.get('【病案介绍】', {}).get('主诉', [])) or None
    except Exception as e:
        description = None
        print(f"提取主诉时发生错误: {e}")

    # 提取年龄、性别
    try:
        case_age, case_sex = (extract_age_and_sex(description) if description else (None, None))
    except Exception as e:
        case_age, case_sex = (None, None)
        print(f"提取年龄和性别时发生错误: {e}")

    # 提取症状
    try:
        case_symptom = case_message.get('【病案介绍】', {}).get('现病史', None)
    except Exception as e:
        case_symptom = None
        print(f"提取症状时发生错误: {e}")

    # 提取既往史
    try:
        case_history = case_message.get('【病案介绍】', {}).get('既往史', None)
    except Exception as e:
        case_history = None
        print(f"提取既往史时发生错误: {e}")

    # 提取体格检查
    try:
        case_physical_exm = case_message.get('【病案介绍】', {}).get('查体', {}).get('体格检查', None)
    except Exception as e:
        case_physical_exm = None
        print(f"提取体格检查时发生错误: {e}")

    # 提取辅助检查
    try:
        case_machine_exm = case_message.get('【病案介绍】', {}).get('查体', {}).get('辅助检查', None)
    except Exception as e:
        case_machine_exm = None
        print(f"提取辅助检查时发生错误: {e}")

    # 提取影像报告
    try:
        case_img_report = case_message.get('【病案介绍】', {}).get('影像报告', None)
    except Exception as e:
        case_img_report = None
        print(f"提取影像报告时发生错误: {e}")

    # 提取诊断结果
    try:
        case_dig_result = case_message.get('tags', {}).get('病种', None)
    except Exception as e:
        case_dig_result = None
        print(f"提取诊断结果时发生错误: {e}")

    # 提取治疗项目
    try:
        case_treatment = case_message.get('【治疗项目】', None)
    except Exception as e:
        case_treatment = None
        print(f"提取治疗项目时发生错误: {e}")

    # 提取一级科室
    try:
        case_first_room = case_message.get('tags', {}).get('科室', None)[0]
    except Exception as e:
        case_first_room = None
        print(f"提取治疗项目时发生错误: {e}")

    # 提取二级科室
    try:
        case_second_room = str(case_message.get('tags', {}).get('科室', None)[1:]).replace("[", "").replace("]", "").replace("'", "")
    except Exception as e:
        case_second_room = None
        print(f"提取治疗项目时发生错误: {e}")

    return case_room, [case_age, case_sex, description, case_symptom, case_history, case_physical_exm, case_machine_exm, case_img_report, case_dig_result, case_treatment, case_first_room, case_second_room]


def cosine_similarity(vec1, vec2):
    # 计算向量的点积
    dot_product = np.dot(vec1, vec2)

    # 计算向量的范数（长度）
    norm_vec1 = np.linalg.norm(vec1)
    norm_vec2 = np.linalg.norm(vec2)

    # 计算余弦相似度
    cosine_sim = dot_product / (norm_vec1 * norm_vec2)

    return cosine_sim


def extract_json_data(sample):
    if "【病案介绍】" not in sample[1].keys() or "主诉" not in sample[1]["【病案介绍】"].keys():
        zhusu = "None"
    else:
        zhusu = ",".join(sample[1]["【病案介绍】"]["主诉"])
    if "【病案介绍】" not in sample[1].keys() or "既往史" not in sample[1]["【病案介绍】"].keys():
        jiwangshi = "None"
    else:
        jiwangshi = " ".join(sample[1]["【病案介绍】"]["既往史"])
    if "【病案介绍】" not in sample[1].keys() or "现病史" not in sample[1]["【病案介绍】"].keys():
        xianbingshi = "None"
    else:
        xianbingshi = " ".join(sample[1]["【病案介绍】"]["现病史"])
    if "【病案介绍】" not in sample[1].keys() or "查体" not in sample[1]["【病案介绍】"].keys():
        chati = "None"
    else:
        chati = sample[1]["【病案介绍】"]["查体"]
    if "tags" not in sample[1].keys() or "科室" not in sample[1]["tags"].keys():
        keshi = "None"
    else:
        keshi = " ".join(sample[1]["tags"]["科室"])
    if "tags" not in sample[1].keys() or "病种" not in sample[1]["tags"].keys():
        jieguo = "None"
    else:
        jieguo = " ".join(sample[1]["tags"]["病种"])
    return zhusu, jiwangshi, xianbingshi, chati, keshi, jieguo



